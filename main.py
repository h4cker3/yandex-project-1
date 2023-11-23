import csv
import io
import os
import re
import sys
from datetime import datetime

from PyQt5 import uic, QtGui
from PyQt5.QtGui import QFont
from PyQt5.QtWidgets import QApplication, QMainWindow, QTableWidget, QTableWidgetItem, \
    QHeaderView, QMessageBox, QAbstractItemView, QAction, QFileDialog, qApp, QInputDialog
from openpyxl import Workbook, load_workbook

# pyuic5 000.uic -o ex.py
# pyqt5-tools designer

# Глобальные переменные
# Настройки приложения
VERSION = "v1.0a"
TODAY = datetime.today().date()
TODAY = TODAY.replace(1, TODAY.month, TODAY.day)
DATA_FORMAT = "[0-3][0-9]\.[0-1][0-9]\.[12][09][0-9][0-9]"
EXIT_CODE_REBOOT = -11725625
# Настройки CSV
DELIMETER = ";"
QUOTECHAR = '"'
# Настройки меню
EXPORT_NAME = "Экспорт в Excel"
IMPORT_NAME = "Импорт из Excel"
SEARCH_NAME = "Поиск по таблице"

# Чтение template из main.ui
template = open('main.ui', 'r', encoding='utf-8').read()


# Ключ для сортировки дат рождений по дню и месяцу (без учета года). На входе получает словарь, соответствующий
# человеку, а возвращает объект date класса datetime, который обладает операторами < и >.
def sort_dates_key(person_dict):
    date = datetime.strptime(person_dict['date'], "%d.%m.%Y").date()
    return date.replace(1, date.month, date.day)


class MainWindow(QMainWindow):
    def __init__(self):
        super(MainWindow, self).__init__()
        f = io.StringIO(template)
        uic.loadUi(f, self)
        self.setWindowIcon(QtGui.QIcon('icon.ico'))

        # Получаем таблицу из UI файла и добавляем необходимые настройки
        self.table = QTableWidget()  # нужно, чтобы Pycharm распознавал класс переменной
        self.table = self.MainTable
        self.table.setColumnCount(3)
        self.table.setHorizontalHeaderLabels(['ФИО', 'Дата рождения', 'Должность'])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)

        # Создаем меню и наполняем его функциями
        self.menubar = self.menuBar()
        self.menu = self.menubar.addMenu("Меню")
        qact = QAction(SEARCH_NAME, self)
        qact.setShortcut('Ctrl+F')
        self.menu.addAction(qact)
        self.menu.addAction(QAction(EXPORT_NAME, self))
        self.menu.addAction(QAction(IMPORT_NAME, self))
        self.menu.triggered[QAction].connect(self.menu_action_trigger)

        # Открываем файл .csv и заполняем таблицу данными
        self.db = open('table.csv', encoding="utf-8")
        data = csv.DictReader(self.db, delimiter=DELIMETER, quotechar=QUOTECHAR)
        self.db_data = sorted(data, key=sort_dates_key)
        self.first_item = 0
        birthday_person = []
        for i, row in enumerate(self.db_data):
            self.table.setRowCount(i + 1)
            for j, value in enumerate(row.values()):
                item = QTableWidgetItem(str(value))
                if sort_dates_key(row) == TODAY and j == 0:
                    if len(birthday_person) == 0:
                        self.first_item = item
                    birthday_person.append([row['full_name'], i])
                self.table.setItem(i, j, item)
        self.table.resizeRowsToContents()
        if len(birthday_person) != 0:
            self.table.scrollToItem(self.first_item, QAbstractItemView.PositionAtTop)
            msg_box = QMessageBox()
            msg_box.setWindowIcon(QtGui.QIcon('icon.ico'))
            msg_box.setIcon(QMessageBox.Information)
            msg_box.setText("Сегодня день рождения у следующих ваших коллег:\n" + "\n".join(
                list(map(lambda x: x[0], birthday_person))))
            msg_box.setWindowTitle("Не забудьте поздравить своих коллег!")
            msg_box.setStandardButtons(QMessageBox.Ok)
            msg_box.exec()
            font = QFont()
            font.setBold(True)
            for person in birthday_person:
                self.table.item(person[1], 0).setBackground(QtGui.QColor(255, 255, 0))
                self.table.item(person[1], 1).setBackground(QtGui.QColor(255, 255, 0))
                self.table.item(person[1], 2).setBackground(QtGui.QColor(255, 255, 0))
                self.table.item(person[1], 0).setFont(font)
                self.table.item(person[1], 1).setFont(font)
                self.table.item(person[1], 2).setFont(font)
        self.db.close()

    def export_to_excel(self):
        path = QFileDialog.getSaveFileName(self, "Выберите файл для экспорта", filter="*.xlsx")
        if not path[0]:
            return
        workbook = Workbook()
        sheet = workbook.active
        sheet['A1'] = "ФИО"
        sheet['B1'] = 'Дата рождения'
        sheet['C1'] = 'Должность'
        for i, row in enumerate(self.db_data):
            sheet[f"A{i + 2}"] = row['full_name']
            sheet[f'B{i + 2}'] = row['date']
            sheet[f'C{i + 2}'] = row['work_place']

        workbook.save(filename=path[0])

    def import_from_excel(self):
        msg_box = QMessageBox()
        msg_box.setIcon(QMessageBox.Information)
        msg_box.setWindowIcon(QtGui.QIcon('icon.ico'))
        msg_box.setText("При импорте из таблицы Excel, данные текущей таблицы будут удалены.\n\nХотите ли вы "
                        "предварительно экспортировать таблицу в Excel?")
        msg_box.setWindowTitle("О сохранении текущих данных")
        msg_box.setStandardButtons(QMessageBox.Yes | QMessageBox.No)
        if msg_box.exec() == QMessageBox.Ok:
            self.export_to_excel()
        path = QFileDialog.getOpenFileName(self, "Выберите файл для импорта", filter="*.xlsx")
        if not path[0]:
            return
        workbook = load_workbook(filename=path[0])
        sheet = workbook.active
        with open('termtable.csv', 'w', newline='', encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=['full_name', 'date', 'work_place'], delimiter=';',
                                    quoting=csv.QUOTE_NONNUMERIC)
            writer.writeheader()
            i = 2
            while sheet.cell(i, 1).value:
                d = {'full_name': sheet.cell(i, 1).value, 'date': sheet.cell(i, 2).value,
                     'work_place': sheet.cell(i, 3).value}
                if not re.match(DATA_FORMAT, d['date']):
                    msg_box = QMessageBox()
                    msg_box.setIcon(QMessageBox.Critical)
                    msg_box.setWindowIcon(QtGui.QIcon('icon.ico'))
                    msg_box.setText(
                        f"Данные в выбранной вами таблице форматированы неправильно.\nПожалуйста, проверьте "
                        f"их и запустите процесс импорта заново.\nОшибка в дате в строке: {i}.")
                    msg_box.setWindowTitle("Ошибка данных!")
                    msg_box.setStandardButtons(QMessageBox.Ok)
                    msg_box.exec()
                    os.remove(os.getcwd() + "\\termtable.csv")
                    return
                writer.writerow(d)
                i += 1
        os.remove(os.getcwd() + "\\table.csv")
        os.rename(os.getcwd() + "\\termtable.csv", "table.csv")
        msg_box = QMessageBox()
        msg_box.setIcon(QMessageBox.Information)
        msg_box.setWindowIcon(QtGui.QIcon('icon.ico'))
        msg_box.setText("Программа будет перезапущена.")
        msg_box.setWindowTitle("Перезапуск программы.")
        msg_box.setStandardButtons(QMessageBox.Ok)
        msg_box.exec()
        qApp.exit(EXIT_CODE_REBOOT)

    def search_in_table(self):
        full_name, res = QInputDialog.getText(
            self, 'Поиск по имени',
            'Введите ФИО своего коллеги, чью дату дня рождения желаете найти:')
        if not res:
            return
        i = 0
        ans = []
        while self.table.item(i, 0):
            if self.table.item(i, 0).text() == full_name.rstrip():
                self.table.scrollToItem(self.table.item(i, 0), QAbstractItemView.PositionAtTop)
                self.table.selectRow(i)
                return
            elif str(self.table.item(i, 0).text()).count(full_name.rstrip()) != 0:
                ans.append([self.table.item(i, 0).text(), self.table.item(i, 1).text()])
            i += 1
        if len(ans) != 0:
            msg_box = QMessageBox()
            msg_box.setIcon(QMessageBox.Information)
            msg_box.setWindowIcon(QtGui.QIcon('icon.ico'))
            str_list = '\n'.join(map(lambda x: x[0] + ' : ' + x[1], ans))
            msg_box.setText(f"Точного совпадения не найдено. Вот подходящие варианты: \n{str_list}")
            msg_box.setWindowTitle("Не точное совпадение.")
            msg_box.setStandardButtons(QMessageBox.Ok)
            msg_box.exec()
            return
        msg_box = QMessageBox()
        msg_box.setIcon(QMessageBox.Warning)
        msg_box.setWindowIcon(QtGui.QIcon('icon.ico'))
        msg_box.setText("Коллега с таким ФИО найден не был.")
        msg_box.setWindowTitle("Не найден.")
        msg_box.setStandardButtons(QMessageBox.Ok)
        msg_box.exec()

    def menu_action_trigger(self, q):
        if q.text() == EXPORT_NAME:
            self.export_to_excel()
        elif q.text() == IMPORT_NAME:
            self.import_from_excel()
        elif q.text() == SEARCH_NAME:
            self.search_in_table()


if __name__ == '__main__':
    app = QApplication(sys.argv)
    currentExitCode = EXIT_CODE_REBOOT
    while currentExitCode == EXIT_CODE_REBOOT:
        try:
            ex = MainWindow()
            ex.setWindowTitle("Дни рождения " + VERSION)
            ex.show()
        except:
            msgbox = QMessageBox()
            msgbox.setIcon(QMessageBox.Critical)
            msgbox.setText(
                "Произошла непредвиденная ошибка в работе программы.\nУдалите файл \"table.csv\", так как он "
                "скорее всего поврежден.")
            msgbox.setWindowTitle("Непредвиденная ошибка!")
            msgbox.setStandardButtons(QMessageBox.Ok)
            msgbox.exec()
            app.exit(0)
        currentExitCode = app.exec()
    sys.exit(currentExitCode)
